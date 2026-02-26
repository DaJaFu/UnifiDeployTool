"""
inventory.py - Collect device info and generate documentation / inventory sheets.
"""

import csv
import json
import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


DEVICE_TYPE_LABELS = {
    "uxg": "UniFi Gateway (UXG)",
    "ugw": "UniFi Security Gateway",
    "usw": "UniFi Switch",
    "uap": "UniFi Access Point",
    "uvc": "UniFi Protect Camera",
    "udm": "UniFi Dream Machine",
}

STATE_LABELS = {
    0: "Pending Adoption",
    1: "Connected",
    2: "Disconnected",
    4: "Upgrading",
    5: "Provisioning",
    6: "Heartbeat Missed",
    7: "Adopting",
    11: "Isolated",
}


def collect_inventory(client) -> dict:
    """Pull all relevant data from the controller and return a structured dict."""
    timestamp = datetime.datetime.now().isoformat()

    try:
        system_info = client.get_system_info()
    except Exception:
        system_info = {}

    try:
        devices_raw = client.get_devices()
    except Exception:
        devices_raw = []

    try:
        networks_raw = client.get_networks()
    except Exception:
        networks_raw = []

    devices = []
    for d in devices_raw:
        dtype = d.get("type", "unknown")
        devices.append({
            "name": d.get("name") or d.get("hostname") or "Unknown",
            "mac": d.get("mac", ""),
            "ip": d.get("ip", ""),
            "model": d.get("model", ""),
            "type": dtype,
            "type_label": DEVICE_TYPE_LABELS.get(dtype, dtype.upper()),
            "firmware": d.get("version", ""),
            "state": d.get("state", -1),
            "state_label": STATE_LABELS.get(d.get("state", -1), "Unknown"),
            "uptime_seconds": d.get("uptime", 0),
            "serial": d.get("serial", ""),
            "site": d.get("site_id", "default"),
        })

    vlans = []
    for n in networks_raw:
        vlans.append({
            "id": n.get("_id", ""),
            "name": n.get("name", ""),
            "vlan": n.get("vlan", ""),
            "purpose": n.get("purpose", ""),
            "subnet": n.get("ip_subnet", ""),
            "dhcp_enabled": n.get("dhcpd_enabled", False),
        })

    gateway_info = {
        "firmware": system_info.get("firmware", {}).get("installed", ""),
        "model": system_info.get("hardware", {}).get("shortname", ""),
        "hostname": system_info.get("hostname", ""),
    }

    return {
        "generated_at": timestamp,
        "gateway": gateway_info,
        "devices": devices,
        "vlans": vlans,
    }


def generate_json(inventory: dict, output_dir: Path) -> Path:
    path = output_dir / "inventory.json"
    with open(path, "w") as f:
        json.dump(inventory, f, indent=2)
    return path


def generate_csv(inventory: dict, output_dir: Path) -> Path:
    path = output_dir / "device_inventory.csv"
    fieldnames = [
        "name", "mac", "ip", "type_label", "model", "firmware",
        "state_label", "serial", "uptime_seconds",
    ]
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(inventory["devices"])
    return path


def generate_xlsx(inventory: dict, output_dir: Path) -> Path | None:
    if not XLSX_AVAILABLE:
        return None

    path = output_dir / "inventory.xlsx"
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Devices ----
    ws_dev = wb.active
    ws_dev.title = "Devices"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)

    dev_headers = ["Name", "MAC", "IP", "Type", "Model", "Firmware", "State", "Serial"]
    for col, h in enumerate(dev_headers, 1):
        cell = ws_dev.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    dev_fields = ["name", "mac", "ip", "type_label", "model", "firmware", "state_label", "serial"]
    for row, dev in enumerate(inventory["devices"], 2):
        for col, field in enumerate(dev_fields, 1):
            ws_dev.cell(row=row, column=col, value=dev.get(field, ""))

    for col in ws_dev.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws_dev.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    # ---- Sheet 2: VLANs ----
    ws_vlan = wb.create_sheet("VLANs")
    vlan_headers = ["Name", "VLAN ID", "Purpose", "Subnet", "DHCP Enabled"]
    for col, h in enumerate(vlan_headers, 1):
        cell = ws_vlan.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    vlan_fields = ["name", "vlan", "purpose", "subnet", "dhcp_enabled"]
    for row, vlan in enumerate(inventory["vlans"], 2):
        for col, field in enumerate(vlan_fields, 1):
            ws_vlan.cell(row=row, column=col, value=str(vlan.get(field, "")))

    # ---- Sheet 3: Summary ----
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "BC10 Deployment Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Generated At", inventory["generated_at"]),
        ("Gateway Model", inventory["gateway"].get("model", "")),
        ("Gateway Firmware", inventory["gateway"].get("firmware", "")),
        ("Total Devices", len(inventory["devices"])),
        ("Total VLANs Configured", len(inventory["vlans"])),
    ]
    for i, (label, value) in enumerate(rows, 3):
        ws_sum.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws_sum.cell(row=i, column=2, value=str(value))

    wb.save(path)
    return path
